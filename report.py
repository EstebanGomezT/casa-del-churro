from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent

IMG_WIDTH_PX = 150
IMG_HEIGHT_PX = 200
ROW_HEIGHT_PT = 155

def create_month_report_xlsx(rows: list[dict], year: int, month: int, date_to_inclusive: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ["Fecha", "Punto Venta", "Total", "Débito", "Crédito", "Efectivo",
               "Bol. Débito", "Bol. Crédito", "Bol. Efectivo", "Folio", "Boleta"]
    ws.append(headers)

    total_sum = debit_sum = credit_sum = cash_sum = 0
    bd_sum = bc_sum = bca_sum = 0

    for idx, r in enumerate(rows):
        ws.append([
            r["date"], r.get("punto_venta", ""),
            r["total"], r["debit"], r["credit"], r["cash"],
            r.get("boletas_debit", 0), r.get("boletas_credit", 0), r.get("boletas_cash", 0),
            r["folio"], ""
        ])
        total_sum += int(r["total"])
        debit_sum += int(r["debit"])
        credit_sum += int(r["credit"])
        cash_sum += int(r["cash"])
        bd_sum += int(r.get("boletas_debit", 0))
        bc_sum += int(r.get("boletas_credit", 0))
        bca_sum += int(r.get("boletas_cash", 0))

        # Insert receipt image
        receipt = r.get("receipt_path", "")
        receipt_file = (BASE_DIR / receipt) if receipt else None
        if receipt_file and receipt_file.exists():
            try:
                img = XlImage(str(receipt_file))
                # Mantener proporcion original
                orig_w, orig_h = img.width, img.height
                if orig_w and orig_h:
                    ratio = min(IMG_WIDTH_PX / orig_w, IMG_HEIGHT_PX / orig_h)
                    img.width = int(orig_w * ratio)
                    img.height = int(orig_h * ratio)
                else:
                    img.width = IMG_WIDTH_PX
                    img.height = IMG_HEIGHT_PX
                cell = f"K{idx + 2}"
                ws.add_image(img, cell)
                ws.row_dimensions[idx + 2].height = ROW_HEIGHT_PT
            except Exception:
                ws.cell(row=idx + 2, column=11, value="(imagen no disponible)")

    ws.append([])
    ws.append(["TOTAL", "", total_sum, debit_sum, credit_sum, cash_sum,
               bd_sum, bc_sum, bca_sum, "", ""])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["K"].width = 24

    # Generar en memoria, sin guardar archivo en disco
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
